from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import pandas as pd
import openpyxl

options = webdriver.ChromeOptions()
options.headless = True
driver = webdriver.Chrome(r'D:\janardhan\WebScrapin\BeauScrap\Drivers\chromedriver.exe', options=options)

#df = pd.read_excel(r'D:\janardhan\WebScrapin\Londonscrap.xlsx', sheet_name='postcodes')
#print(df)

wb_obj = openpyxl.load_workbook(r'D:\janardhan\WebScrapin\Londonscrap.xlsx')
sheet_obj1 = wb_obj["postcodes"]
sheet_obj2 = wb_obj["services"]
m_rowsh1 = sheet_obj1.max_row
m_colsh1 = sheet_obj1.max_column
m_rowsh2 = sheet_obj2.max_row
m_colsh2 = sheet_obj2.max_column


#Working scraping
for i in range(1, 1):
    for h in range(1, m_colsh1 + 1):
        cell_objsh1 = sheet_obj1.cell(row=i, column=h)
        for j in range(1, 1 + 1):
            for k in range(1, m_colsh2 + 1):
                cell_objsh2 = sheet_obj2.cell(row=j, column=k)
                print(cell_objsh1.value)
                print(cell_objsh2.value)
                driver.get(f"https://www.google.com/maps/search/{cell_objsh1.value}+{cell_objsh2.value}")
                element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//div[contains(@aria-label,'Results for')]")))
                eleContent = driver.find_elements_by_xpath("//div[@class='section-result-text-content']")
                thislist = list()
                emptyDict = dict()

                for k in eleContent:
                    my_tuple = (k.text.splitlines()[0], k.text.splitlines()[len(k.text.splitlines()) - 1])
                    thislist.append(my_tuple)
                    print(thislist)
                    print(len(thislist))
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    for i in range(len(thislist)):
                        sheet.append(thislist[i])
                    wb.save(f"D:\janardhan\WebScrapin\demo.xlsx")